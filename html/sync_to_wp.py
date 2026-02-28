#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WordPress 产品同步脚本
从 Excel 读取产品数据，通过 WP REST API 同步到 WordPress
"""

import json
import time
import os
import sys
import requests
from urllib.parse import urlparse
from openpyxl import load_workbook

# ============ 配置 ============
WP_URL = "https://zr.jsss999.com/wp-json/wp/v2"
WP_USER = "*"
WP_APP_PASSWORD = "*"
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zgz_products_import.xlsx")
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sync_result.log")
PROGRESS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sync_progress.json")
POST_STATUS = "draft"  # draft=草稿, publish=直接发布
SLEEP_INTERVAL = 1     # 每条之间间隔秒数

# ============ 分类映射 ============
# CSV category_path -> WP 二级分类 ID
CATEGORY_MAP = {
    "加热器->导热油加热器": 42,
    "加热器->管道加热器系列产品": 43,
    "加热器->风道加热器系列产品": 44,
    "加热器->电加热器系列产品": 45,
    "加热器->电加热元件系列产品": 46,
    "搅拌设备->搅拌罐系列产品": 47,
    "搅拌设备->搅拌器系列产品": 48,
    "水处理设备->过滤器系列产品": 49,
    "水处理设备->除污器系列产品": 50,
    "水处理设备->分集水器系列产品": 51,
}

# ============ 标签映射 ============
# CSV category_path -> WP 标签 ID 列表（每篇文章都带"中热"标签 + 对应一级分类标签）
TAG_MAP = {
    "加热器->导热油加热器": [40, 63],
    "加热器->管道加热器系列产品": [40, 63],
    "加热器->风道加热器系列产品": [40, 63],
    "加热器->电加热器系列产品": [40, 63],
    "加热器->电加热元件系列产品": [40, 63],
    "搅拌设备->搅拌罐系列产品": [40, 64],
    "搅拌设备->搅拌器系列产品": [40, 64],
    "水处理设备->过滤器系列产品": [40, 65],
    "水处理设备->除污器系列产品": [40, 65],
    "水处理设备->分集水器系列产品": [40, 65],
}


def log(msg):
    """打印并写入日志"""
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(msg + "\n")


def load_progress():
    """加载已同步进度，用于断点续传"""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_progress(progress):
    """保存同步进度"""
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def make_progress_key(product_id, category_path):
    """生成去重 key: product_id + 分类"""
    return f"{product_id}_{category_path}"


def read_excel():
    """读取 Excel 文件，返回产品列表"""
    products = []
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active

    # 获取表头
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    # 读取数据行
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_data[headers[i]] = str(cell.value) if cell.value is not None else ""

        cat = row_data.get("category_path", "").strip()
        # 跳过无效分类
        if "->" not in cat:
            continue
        if cat not in CATEGORY_MAP:
            log(f"[跳过] 未知分类: {cat}, product_id: {row_data.get('product_id')}")
            continue
        products.append(row_data)

    wb.close()
    return products


def upload_featured_image(image_url, title):
    """下载图片并上传到 WP 媒体库，返回 media_id"""
    try:
        # 下载图片
        resp = requests.get(image_url, timeout=30)
        if resp.status_code != 200:
            log(f"  [图片下载失败] {image_url} status={resp.status_code}")
            return 0

        # 从 URL 提取文件名
        parsed = urlparse(image_url)
        filename = os.path.basename(parsed.path)
        if not filename:
            filename = "product.jpg"

        # 判断 content-type
        content_type = resp.headers.get("Content-Type", "image/jpeg")

        # 上传到 WP
        headers = {
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": content_type,
        }
        wp_resp = requests.post(
            f"{WP_URL}/media",
            auth=(WP_USER, WP_APP_PASSWORD),
            headers=headers,
            data=resp.content,
            timeout=60,
        )

        if wp_resp.status_code in (200, 201):
            media_id = wp_resp.json().get("id", 0)
            log(f"  [图片上传成功] media_id={media_id}")
            return media_id
        else:
            log(f"  [图片上传失败] status={wp_resp.status_code} {wp_resp.text[:200]}")
            return 0

    except Exception as e:
        log(f"  [图片异常] {e}")
        return 0


def create_post(title, content, category_id, tag_ids, featured_media_id):
    """创建 WP 文章，返回 post_id"""
    data = {
        "title": title,
        "content": content,
        "categories": [category_id],
        "tags": tag_ids,
        "status": POST_STATUS,
    }
    if featured_media_id > 0:
        data["featured_media"] = featured_media_id

    try:
        resp = requests.post(
            f"{WP_URL}/posts",
            auth=(WP_USER, WP_APP_PASSWORD),
            json=data,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            post_id = resp.json().get("id", 0)
            return post_id
        else:
            log(f"  [文章创建失败] status={resp.status_code} {resp.text[:200]}")
            return 0
    except Exception as e:
        log(f"  [文章异常] {e}")
        return 0


def sync_products(products, force=False):
    """同步产品列表"""
    progress = load_progress()

    success = 0
    skip = 0
    fail = 0

    for i, row in enumerate(products):
        product_id = row["product_id"].strip()
        title = row["title"].strip()
        category_path = row["category_path"].strip()
        content_html = row["content_html"].strip()
        oss_images = row["oss_image_url"].strip()

        # 去重检查
        key = make_progress_key(product_id, category_path)
        if key in progress and not force:
            existing = progress.get(key) or {}
            existing_title = existing.get("title") or title
            existing_post_id = existing.get("post_id")
            existing_media_id = existing.get("media_id")
            log(
                "  [跳过] 远程已采集/本地已记录: "
                f"product_id={product_id} 分类={category_path} 标题={existing_title} "
                f"post_id={existing_post_id} media_id={existing_media_id}"
            )
            skip += 1
            continue

        category_id = CATEGORY_MAP[category_path]
        tag_ids = TAG_MAP.get(category_path, [40])
        first_image = oss_images.split("|")[0].strip() if oss_images else ""

        log(f"[{i+1}/{len(products)}] product_id={product_id} 标题={title} 分类={category_path}")

        # 上传特色图片
        media_id = 0
        if first_image:
            media_id = upload_featured_image(first_image, title)

        # 创建文章
        post_id = create_post(title, content_html, category_id, tag_ids, media_id)

        if post_id > 0:
            log(f"  [成功] post_id={post_id}\n")
            progress[key] = {
                "post_id": post_id,
                "media_id": media_id,
                "product_id": product_id,
                "title": title,
            }
            save_progress(progress)
            success += 1
        else:
            log(f"  [失败]\n")
            fail += 1

        time.sleep(SLEEP_INTERVAL)

    return success, skip, fail


def main():
    import argparse
    parser = argparse.ArgumentParser(description="WordPress 产品同步脚本")
    parser.add_argument("--id", type=str, help="只同步指定 product_id（单条测试）")
    parser.add_argument("--force", action="store_true", help="忽略去重，强制重新同步")
    parser.add_argument("--publish", action="store_true", help="直接发布（默认草稿）")
    args = parser.parse_args()

    global POST_STATUS
    if args.publish:
        POST_STATUS = "publish"

    log("=" * 60)
    log(f"开始同步 - 状态: {POST_STATUS}")
    log("=" * 60)

    products = read_excel()
    log(f"Excel 共 {len(products)} 条有效产品\n")

    # 如果指定了 --id，只同步该条
    if args.id:
        products = [p for p in products if p["product_id"].strip() == args.id]
        if not products:
            log(f"未找到 product_id={args.id}")
            return
        log(f"单条模式: product_id={args.id}\n")

    progress = load_progress()
    log(f"已同步 {len(progress)} 条，将跳过\n")

    success, skip, fail = sync_products(products, force=args.force)

    log("=" * 60)
    log(f"同步完成: 成功 {success}, 跳过 {skip}, 失败 {fail}")
    log(f"总计: {success + skip + fail} / {len(products)}")
    log("=" * 60)


if __name__ == "__main__":
    main()
