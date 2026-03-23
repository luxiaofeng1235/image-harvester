"""将4个Excel导入 jj_wangyi_goods 表"""
import sys
import time
from datetime import datetime
from pathlib import Path

import pymysql
from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parent))

from ai_goods_pipeline.config import load_settings
from ai_goods_pipeline.enums.source_types import SOURCE_LEGACY_IMPORT
from ai_goods_pipeline.utils.batch_meta import build_source_note, normalize_batch_id

FILES = [
    (126, "goods_苏州特产.xlsx"),
    (127, "goods_农副产品.xlsx"),
    (128, "goods_苏超纪念品.xlsx"),
    (129, "goods_工艺产品.xlsx"),
]

SQL = """INSERT INTO jj_wangyi_goods
    (
        goods_name, sub_title, category_id, image, price, description,
        batch_id, last_batch_id, source_type, source_note, create_time, update_time
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""


def main():
    settings = load_settings()
    conn = pymysql.connect(
        host=settings.db_host,
        port=settings.db_port,
        user=settings.db_user,
        password=settings.db_password,
        database=settings.db_name,
        charset=settings.db_charset,
    )
    cursor = conn.cursor()
    now = int(time.time())
    batch_id = normalize_batch_id(
        datetime.now().strftime("%Y%m%d_%H%M%S"),
        fallback="legacy_import",
    )
    total = 0

    for cat_id, fname in FILES:
        wb = load_workbook(fname)
        ws = wb.active
        count = 0
        source_note = build_source_note([f"excel={fname}"])
        for row in range(2, ws.max_row + 1):
            title = ws.cell(row, 2).value or ""
            image = ws.cell(row, 3).value or ""
            price = ws.cell(row, 4).value or 0
            subtitle = ws.cell(row, 5).value or ""
            desc = ws.cell(row, 6).value or ""

            cursor.execute(
                SQL,
                (
                    title,
                    subtitle,
                    cat_id,
                    image,
                    price,
                    desc,
                    batch_id,
                    batch_id,
                    SOURCE_LEGACY_IMPORT,
                    source_note,
                    now,
                    now,
                ),
            )
            count += 1

        conn.commit()
        total += count
        print(f"分类{cat_id} | {fname} | 导入 {count} 条 | batch_id={batch_id}")

    cursor.close()
    conn.close()
    print(f"\n完成，共导入 {total} 条 | batch_id={batch_id}")


if __name__ == "__main__":
    main()
